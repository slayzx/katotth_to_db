import re, os
import psycopg2, psycopg2.errors
# from icecream import ic
from openpyxl import load_workbook
from dotenv import load_dotenv, dotenv_values

config = dotenv_values(".env")

load_dotenv()

# crimea = re.compile('республіка крим', re.I)

object_decode = {
    'O': 'область',
    'K': 'місто зі спеціальним статусом',  # city with special status
    'P': 'район',
    'H': 'територіальна громада',
    'M': 'місто',
    'T': 'смт',
    'C': 'село',
    'X': 'селище',
    'B': 'район у місті'
}
hromada_type_decode = {
    'M': 'міська',
    'T': 'селищна',
    'X': 'сільська',
    'C': 'сільська',
}
object_to_column = {
    'O': 0,
    'K': 0,  # city with special status
    'P': 1,
    'H': 2,
    'M': 3,
    'T': 3,
    'C': 3,
    'X': 3,
    'B': 4
}

'''Columns: A - region, B - district, C - hromada, D - communities, 
E - districts in cities, F - object category, G - name'''

# variables for database: reg - region, dist - district, hrom - hromada, munic - municipalities,
#  division_name - unique number of administrative division object, div_type - type of division object
# respectively to katotth, division_name - name of administrative division object,
# division_full_name - humanised division_name + div_type

past_district, spec_city = 0, False

region, district, district_name, hromada, municipal, district_city, katotth, division_type, division_name, \
division_full_name = 0, 0, '', 0, 0, 0, 0, '', '', ''

conn = psycopg2.connect(
    host=config['PGSQL_HOST'],
    database=config['PGSQL_DB_NAME'],
    user=config['PGSQL_DB_USER'],
    password='PGSQL_DB_PASSWD')

drop_table = """DROP TABLE katotth;"""

create_table = """create table katotth( 
                id SERIAL PRIMARY KEY, 
                region SMALLINT, 
                region_name VARCHAR(50),
                distr SMALLINT, 
                distr_name VARCHAR(50),
                hrom SMALLINT, 
                hrom_name VARCHAR(100),
                municip SMALLINT, 
                municip_name VARCHAR(50),
                distr_city SMALLINT, 
                katotth VARCHAR(20), 
                div_type VARCHAR(50), 
                div_name VARCHAR(50), 
                div_full_name VARCHAR(100) 
                );"""


def exel_files_list() -> list:
    exel_files = []
    for f_name in os.listdir('./'):
        if f_name.endswith('.xlsx'):
            exel_files.append(f_name)
    return exel_files


def find_upper_left_cell(workbook_sheet):
    for i in range(1, 8):
        upper_cell = workbook_sheet.cell(row=i, column=1)
        if upper_cell.value is None:
            continue
        elif re.match('(UA\\d{17})', upper_cell.value):
            return upper_cell.coordinate


def find_lower_right_cell(workbook_sheet):
    bottom_cell = ''
    mx_row = workbook_sheet.max_row
    mx_col = workbook_sheet.max_column
    # bottom_row = ''
    # bottom_cell = ''
    for i in range(mx_row, mx_row - 6, -1):
        cell_for_row = workbook_sheet.cell(row=i, column=1)
        if cell_for_row.value is None:
            continue
        elif re.match('(UA\\d{17})', cell_for_row.value):
            bottom_row = cell_for_row.row
            for j in range(1, mx_col + 2):
                cell_for_column = workbook_sheet.cell(row=bottom_row, column=j)
                if cell_for_column.value is None:
                    break
                else:
                    bottom_cell = cell_for_column.coordinate
                    continue
            break
    return bottom_cell


try:
    with conn:
        with conn.cursor() as curs:
            curs.execute(create_table)
except psycopg2.errors.DuplicateTable:
    with conn:
        with conn.cursor() as curs:
            curs.execute(drop_table)
            curs.execute(create_table)
conn.commit()

files_list = sorted(exel_files_list())
files_count = len(files_list)

if not files_count:
    print('Put the Exel file in root directory')
    answer = input("If you have put the file enter 'y' and press enter key")
    while answer not in ['y', 'n']:
        answer = input('.. so we continue? y/n')
    if answer == 'n':
        print('See you soon')
        exit()
    files_list = sorted(exel_files_list())
    files_count = len(files_list)
elif files_count == 1:
    file = files_list[0]
    print(f"Завантажується файл '{file}'...")
    workbook = load_workbook(file)
    print(f"Завантажено файл '{file}'")
elif files_count > 1:
    print(f'В папці більше одного Exel файлу, виберіть номер, та введіть його:')
    for f in range(1, files_count + 1):
        print(f'{f} : {files_list[f - 1]}')
    choose = int(input()) - 1
    file = files_list[choose]
    print(f"Завантажується файл '{file}'...")
    workbook = load_workbook(file)
    print(f"Завантажено файл '{file}'")

sheet = workbook.active

upper_left = find_upper_left_cell(sheet)
lower_right = find_lower_right_cell(sheet)
print('Перевірка літер типу адмін одиниці...')
types_list = object_decode.keys()
errors_in_types = False
for i in range(int(upper_left[1:]), int(lower_right[1:]) + 1):
    c = sheet.cell(row=i, column=6)
    if c.value not in types_list:
        errors_in_types = True
        print('Тип має невірне значення: ', c.value, ' -> ', c.coordinate)
if errors_in_types:
    print(f"Виправте вказані вище помилки у файлі '{file}' , та спробуйте знову")
    exit()

with conn:
    with conn.cursor() as curs:

        for row in sheet[upper_left:lower_right]:
            division_type = row[5].value.strip()
            division_name = row[6].value.strip()
            if division_type == 'O':  # region
                if re.search('республіка крим', division_name, re.I):
                    division_full_name = division_name
                else:
                    division_full_name = f"{division_name} {object_decode[division_type]}"
                print(f"Обробляється {division_full_name}")
            elif division_type in ['P']:  # district
                division_full_name = f"{division_name} {object_decode[division_type]}"
            elif division_type in ['H']:  # hromada
                next_row = row[5].row + 1
                next_row_cell = sheet.cell(row=next_row, column=6)
                hromada_type = next_row_cell.value
                division_full_name = f"{division_name} {hromada_type_decode[hromada_type]} {object_decode[division_type]}"
                pass
            elif division_type == 'B':  # district in city
                division_full_name = f"{division_name} {object_decode[division_type]}"
            else:
                division_full_name = f"{object_decode[division_type]} {division_name}"

            past_region, past_district, past_hromada, = region, district, hromada

            atu_num = row[object_to_column[division_type]].value

            region = int(atu_num[2:4])
            district = int(atu_num[4:6])
            hromada = int(atu_num[6:9])
            municipal = int(atu_num[9:12])
            district_city = int(atu_num[12:14])
            katotth = atu_num

            # If object of administrative division is region or city with special status
            if division_type == 'O':
                if past_region != region:
                    region_name = division_full_name
                curs.execute(
                    """INSERT INTO katotth(region, region_name, distr, hrom, municip, katotth, div_type,
                                           div_name, div_full_name) 
                       VALUES(%s, %s, %s, %s, %s, %s, %s, %s, %s)""",
                    (region, region_name, district, hromada, municipal, katotth,
                     object_decode[division_type], division_name, division_full_name))
            # If object of administrative division is city with spec status
            elif division_type == 'K':
                spec_city = True
                division_full_name = f"місто {division_name}"
                if past_municipal != municipal:
                    past_municipal = division_name
                curs.execute(
                    """INSERT INTO katotth(region, region_name, municip, katotth, div_type,
                                           div_name, div_full_name)
                       VALUES(%s, %s, %s, %s, %s, %s, %s)""",
                    (region, region_name, municipal, katotth,
                     object_decode[division_type], division_name, division_full_name))
            # If object of administrative division is district
            elif division_type == 'P':
                if past_district != district:
                    district_name = division_full_name
                curs.execute(
                    """INSERT INTO katotth(region, region_name, distr, distr_name, hrom, municip, katotth,
                                           div_type, div_name, div_full_name) 
                       VALUES(%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)""",
                    (region, region_name, district, district_name, hromada, municipal, katotth,
                     object_decode[division_type], division_name, division_full_name))
            # If object of administrative division is territorial hromada
            elif division_type == 'H':
                if past_hromada != hromada:
                    hromada_name = division_full_name
                curs.execute(
                    """INSERT INTO katotth(region, region_name, distr, distr_name, hrom, hrom_name, municip, 
                                           katotth, div_type, div_name, div_full_name) 
                       VALUES(%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)""",
                    (region, region_name, district, district_name, hromada,
                     hromada_name, municipal,
                     katotth, object_decode[division_type], division_name, division_full_name))
            # If object of administrative division is one of the city or or village or special status
            elif division_type in ['M', 'T', 'C', 'X']:
                past_municipal = division_name
                curs.execute(
                    """INSERT INTO katotth(region, region_name, distr, distr_name, hrom, hrom_name, municip, katotth, 
                    div_type, div_name, div_full_name) 
                       VALUES(%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)""",
                    (region, region_name, district, district_name, hromada, hromada_name, municipal, katotth,
                     object_decode[division_type], division_name, division_full_name))
            # If object of administrative division is district in city
            elif division_type == 'B':
                if spec_city:
                    region_name = past_municipal
                    curs.execute(
                        """INSERT INTO katotth(region, region_name, municip_name, distr_city, katotth, div_type, 
                        div_name, div_full_name) 
                            VALUES(%s, %s, %s, %s, %s, %s, %s, %s)""",
                        (region, region_name, past_municipal, district_city, katotth, object_decode[division_type],
                         division_name, f"{division_full_name} {past_municipal}"))
                else:
                    curs.execute(
                        """INSERT INTO katotth(region, region_name, distr, distr_name, hrom, hrom_name, municip, 
                        municip_name, distr_city, katotth, div_type, div_name, div_full_name) 
                            VALUES(%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)""",
                        (region, region_name, district, district_name, hromada, hromada_name, municipal, past_municipal,
                         district_city, katotth, object_decode[division_type], division_name,
                         f"{division_full_name} {past_municipal}"))
    conn.commit()
    pass

workbook.close()
